from __future__ import annotations

import json
import os
from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import pandas as pd
from sklearn.base import clone
from sklearn.compose import ColumnTransformer
from sklearn.ensemble import GradientBoostingClassifier, RandomForestClassifier
from sklearn.impute import SimpleImputer
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import (
    accuracy_score,
    balanced_accuracy_score,
    confusion_matrix,
    f1_score,
    log_loss,
    roc_auc_score,
    roc_curve,
)
from sklearn.model_selection import StratifiedKFold
from sklearn.pipeline import Pipeline
from sklearn.preprocessing import LabelBinarizer, LabelEncoder, OneHotEncoder, StandardScaler
from sklearn.svm import SVC

try:
    from xgboost import XGBClassifier
except Exception:  # pragma: no cover
    XGBClassifier = None


ROOT = Path(__file__).resolve().parents[1]
DATASET = ROOT / "dataset.csv"
DATA_DIR = ROOT / "data"
OUTPUT = ROOT / "output" / "pcs_symptom_task"
RANDOM_STATE = 42
N_SPLITS = 5


RAW_TO_LABEL = {
    "腹痛": "abdominal_pain",
    "便秘": "constipation",
    "胃肠功能紊乱": "gastrointestinal_dysfunction",
    "消化不良": "dyspepsia",
    "肠道菌群失调": "intestinal_dysbiosis",
    "慢性胃炎": "chronic_gastritis",
    "腹胀": "abdominal_distension",
}

LABEL_TO_CATEGORY = {
    "abdominal_pain": "pain",
    "constipation": "constipation",
    "gastrointestinal_dysfunction": "gi_dysfunction_dyspepsia",
    "dyspepsia": "gi_dysfunction_dyspepsia",
    "intestinal_dysbiosis": "gi_dysfunction_dyspepsia",
    "chronic_gastritis": "gi_dysfunction_dyspepsia",
    "abdominal_distension": "gi_dysfunction_dyspepsia",
}

CATEGORY_DISPLAY = {
    "no_pcs": "No PCS",
    "pain": "Pain phenotype",
    "constipation": "Constipation phenotype",
    "gi_dysfunction_dyspepsia": "GI dysfunction / dyspepsia phenotype",
}


def make_one_hot_encoder() -> OneHotEncoder:
    try:
        return OneHotEncoder(handle_unknown="ignore", sparse_output=False)
    except TypeError:
        return OneHotEncoder(handle_unknown="ignore", sparse=False)


def source_xlsx() -> Path:
    files = [DATA_DIR / name for name in os.listdir(DATA_DIR) if name.endswith(".xlsx")]
    if not files:
        raise FileNotFoundError("No .xlsx file found in data directory")
    return files[0]


def extract_symptom_text() -> pd.DataFrame:
    workbook = openpyxl.load_workbook(source_xlsx(), read_only=True, data_only=True)
    sheet = workbook.active
    rows = []
    for row_id in range(3, sheet.max_row + 1):
        rows.append(
            {
                "source_row": row_id,
                "pcs_raw": sheet.cell(row_id, 39).value,
                "pcs_onset_time_raw": sheet.cell(row_id, 40).value,
                "pcs_symptom_raw": sheet.cell(row_id, 41).value,
            }
        )
    workbook.close()
    return pd.DataFrame(rows)


def normalize_symptom_label(row: pd.Series) -> tuple[str, str]:
    raw = row["pcs_symptom_raw"]
    pcs = int(row["pcs"])
    if pcs == 0:
        return "no_pcs", "no_pcs"
    if pd.isna(raw) or raw == "":
        return "pcs_unknown", "pcs_unknown"
    label = RAW_TO_LABEL.get(str(raw).strip(), "other_pcs_symptom")
    category = LABEL_TO_CATEGORY.get(label, "gi_dysfunction_dyspepsia")
    return label, category


def build_symptom_dataset() -> pd.DataFrame:
    base = pd.read_csv(DATASET)
    symptoms = extract_symptom_text()
    if len(base) != len(symptoms):
        raise ValueError(f"Dataset rows ({len(base)}) do not match source symptom rows ({len(symptoms)})")
    out = base.copy()
    out.insert(0, "source_row", symptoms["source_row"].to_numpy())
    out["pcs_onset_time_raw"] = symptoms["pcs_onset_time_raw"]
    out["pcs_symptom_raw"] = symptoms["pcs_symptom_raw"]
    labels = out.apply(normalize_symptom_label, axis=1, result_type="expand")
    out["pcs_symptom_label"] = labels[0]
    out["pcs_symptom_category"] = labels[1]
    return out


def build_preprocessor(x: pd.DataFrame) -> ColumnTransformer:
    categorical_cols = x.select_dtypes(include=["object", "category"]).columns.tolist()
    numeric_cols = [col for col in x.columns if col not in categorical_cols]
    numeric_pipeline = Pipeline(
        steps=[
            ("imputer", SimpleImputer(strategy="median")),
            ("scaler", StandardScaler()),
        ]
    )
    categorical_pipeline = Pipeline(
        steps=[
            ("imputer", SimpleImputer(strategy="most_frequent")),
            ("onehot", make_one_hot_encoder()),
        ]
    )
    return ColumnTransformer(
        transformers=[
            ("numeric", numeric_pipeline, numeric_cols),
            ("categorical", categorical_pipeline, categorical_cols),
        ]
    )


def build_models(n_classes: int, class_weights: dict[int, float]) -> dict[str, object]:
    models: dict[str, object] = {
        "Multinomial Logistic": LogisticRegression(
            max_iter=5000,
            class_weight="balanced",
            solver="lbfgs",
            multi_class="auto",
            random_state=RANDOM_STATE,
        ),
        "Random Forest": RandomForestClassifier(
            n_estimators=500,
            max_depth=6,
            min_samples_leaf=3,
            class_weight="balanced_subsample",
            random_state=RANDOM_STATE,
            n_jobs=-1,
        ),
        "Gradient Boosting": GradientBoostingClassifier(
            n_estimators=180,
            learning_rate=0.04,
            max_depth=2,
            random_state=RANDOM_STATE,
        ),
        "SVM": SVC(
            kernel="rbf",
            C=1.0,
            gamma="scale",
            class_weight="balanced",
            probability=True,
            random_state=RANDOM_STATE,
        ),
    }
    if XGBClassifier is not None:
        models["XGBoost"] = XGBClassifier(
            n_estimators=220,
            max_depth=2,
            learning_rate=0.04,
            subsample=0.85,
            colsample_bytree=0.85,
            reg_lambda=2.0,
            objective="multi:softprob",
            num_class=n_classes,
            eval_metric="mlogloss",
            random_state=RANDOM_STATE,
            n_jobs=-1,
        )
    return models


def class_weight_map(y_encoded: np.ndarray) -> dict[int, float]:
    classes, counts = np.unique(y_encoded, return_counts=True)
    total = len(y_encoded)
    return {int(cls): total / (len(classes) * count) for cls, count in zip(classes, counts)}


def multiclass_auc(y_true: np.ndarray, prob: np.ndarray, labels: list[int]) -> float:
    try:
        return roc_auc_score(y_true, prob, labels=labels, multi_class="ovr", average="macro")
    except ValueError:
        return np.nan


def normalize_probabilities(prob: np.ndarray) -> np.ndarray:
    prob = np.asarray(prob, dtype=float)
    row_sums = prob.sum(axis=1, keepdims=True)
    row_sums[row_sums == 0] = 1.0
    return prob / row_sums


def evaluate_symptom_task(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, LabelEncoder]:
    model_df = df.loc[df["pcs_symptom_category"] != "pcs_unknown"].copy()
    y_text = model_df["pcs_symptom_category"]
    predictors = model_df.drop(
        columns=[
            "source_row",
            "pcs",
            "pcs_onset_time_raw",
            "pcs_symptom_raw",
            "pcs_symptom_label",
            "pcs_symptom_category",
        ]
    )
    label_encoder = LabelEncoder()
    y = label_encoder.fit_transform(y_text)
    labels = list(range(len(label_encoder.classes_)))
    cv = StratifiedKFold(n_splits=N_SPLITS, shuffle=True, random_state=RANDOM_STATE)
    preprocessor = build_preprocessor(predictors)
    models = build_models(len(labels), class_weight_map(y))

    fold_rows = []
    oof_rows = []
    summary_rows = []
    probability_store: dict[str, np.ndarray] = {}

    for model_name, estimator in models.items():
        oof_prob = np.zeros((len(model_df), len(labels)))
        oof_pred = np.zeros(len(model_df), dtype=int)
        for fold, (train_idx, test_idx) in enumerate(cv.split(predictors, y), start=1):
            pipeline = Pipeline(
                steps=[
                    ("preprocess", clone(preprocessor)),
                    ("model", clone(estimator)),
                ]
            )
            pipeline.fit(predictors.iloc[train_idx], y[train_idx])
            prob = normalize_probabilities(pipeline.predict_proba(predictors.iloc[test_idx]))
            pred = np.argmax(prob, axis=1)
            oof_prob[test_idx] = prob
            oof_pred[test_idx] = pred
            fold_rows.append(
                {
                    "model": model_name,
                    "fold": fold,
                    "accuracy": accuracy_score(y[test_idx], pred),
                    "balanced_accuracy": balanced_accuracy_score(y[test_idx], pred),
                    "macro_f1": f1_score(y[test_idx], pred, average="macro", zero_division=0),
                    "weighted_f1": f1_score(y[test_idx], pred, average="weighted", zero_division=0),
                    "macro_ovr_auc": multiclass_auc(y[test_idx], prob, labels),
                    "log_loss": log_loss(y[test_idx], prob, labels=labels),
                }
            )

        probability_store[model_name] = oof_prob
        for idx, row_id in enumerate(model_df["source_row"]):
            row = {
                "source_row": int(row_id),
                "observed_category": label_encoder.inverse_transform([y[idx]])[0],
                "predicted_category": label_encoder.inverse_transform([oof_pred[idx]])[0],
                "model": model_name,
            }
            for class_idx, class_name in enumerate(label_encoder.classes_):
                row[f"prob_{class_name}"] = oof_prob[idx, class_idx]
            oof_rows.append(row)

        summary_rows.append(
            {
                "model": model_name,
                "oof_accuracy": accuracy_score(y, oof_pred),
                "oof_balanced_accuracy": balanced_accuracy_score(y, oof_pred),
                "oof_macro_f1": f1_score(y, oof_pred, average="macro", zero_division=0),
                "oof_weighted_f1": f1_score(y, oof_pred, average="weighted", zero_division=0),
                "oof_macro_ovr_auc": multiclass_auc(y, oof_prob, labels),
                "oof_log_loss": log_loss(y, oof_prob, labels=labels),
            }
        )

    fold_df = pd.DataFrame(fold_rows)
    summary_df = pd.DataFrame(summary_rows).sort_values(
        ["oof_macro_f1", "oof_macro_ovr_auc"], ascending=[False, False]
    )
    oof_df = pd.DataFrame(oof_rows)
    plot_outputs(model_df, y, label_encoder, summary_df.iloc[0]["model"], probability_store[summary_df.iloc[0]["model"]])
    return fold_df, summary_df, oof_df, label_encoder


def configure_plots() -> None:
    plt.rcParams.update(
        {
            "font.family": "Arial",
            "figure.dpi": 150,
            "savefig.dpi": 300,
            "axes.spines.top": False,
            "axes.spines.right": False,
            "axes.grid": True,
            "grid.alpha": 0.25,
        }
    )


def plot_outputs(model_df: pd.DataFrame, y: np.ndarray, encoder: LabelEncoder, best_model: str, best_prob: np.ndarray) -> None:
    configure_plots()
    counts = model_df["pcs_symptom_category"].value_counts().reindex(encoder.classes_)
    labels = [CATEGORY_DISPLAY.get(cls, cls) for cls in counts.index]
    plt.figure(figsize=(7.4, 4.6))
    bars = plt.bar(labels, counts.values, color=["#4c78a8", "#e45756", "#72b7b2", "#f2b447"][: len(labels)])
    for bar, value in zip(bars, counts.values):
        plt.text(bar.get_x() + bar.get_width() / 2, bar.get_height(), str(int(value)), ha="center", va="bottom")
    plt.ylabel("Number of patients")
    plt.title("PCS symptom occurrence classes")
    plt.xticks(rotation=25, ha="right")
    plt.tight_layout()
    plt.savefig(OUTPUT / "pcs_symptom_class_distribution.png", bbox_inches="tight")
    plt.close()

    pred = np.argmax(best_prob, axis=1)
    cm = confusion_matrix(y, pred, labels=list(range(len(encoder.classes_))))
    plt.figure(figsize=(6.4, 5.4))
    plt.imshow(cm, cmap="Blues")
    plt.colorbar(fraction=0.046, pad=0.04)
    display = [CATEGORY_DISPLAY.get(cls, cls) for cls in encoder.classes_]
    plt.xticks(np.arange(len(display)), display, rotation=35, ha="right")
    plt.yticks(np.arange(len(display)), display)
    for i in range(cm.shape[0]):
        for j in range(cm.shape[1]):
            plt.text(j, i, str(cm[i, j]), ha="center", va="center", color="#111827")
    plt.xlabel("Predicted")
    plt.ylabel("Observed")
    plt.title(f"OOF confusion matrix: {best_model}")
    plt.tight_layout()
    plt.savefig(OUTPUT / "pcs_symptom_confusion_matrix.png", bbox_inches="tight")
    plt.close()

    lb = LabelBinarizer()
    y_bin = lb.fit_transform(y)
    plt.figure(figsize=(7.2, 5.4))
    for class_idx, class_name in enumerate(encoder.classes_):
        fpr, tpr, _ = roc_curve(y_bin[:, class_idx], best_prob[:, class_idx])
        auc = roc_auc_score(y_bin[:, class_idx], best_prob[:, class_idx])
        plt.plot(fpr, tpr, lw=2, label=f"{CATEGORY_DISPLAY.get(class_name, class_name)} (AUC={auc:.3f})")
    plt.plot([0, 1], [0, 1], linestyle="--", color="#8a8f99", lw=1.2)
    plt.xlabel("False positive rate")
    plt.ylabel("True positive rate")
    plt.title(f"One-vs-rest ROC curves: {best_model}")
    plt.legend(loc="lower right", fontsize=8)
    plt.tight_layout()
    plt.savefig(OUTPUT / "pcs_symptom_ovr_roc.png", bbox_inches="tight")
    plt.close()


def markdown_table(df: pd.DataFrame) -> str:
    columns = df.columns.tolist()
    lines = [
        "| " + " | ".join(columns) + " |",
        "| " + " | ".join(["---"] * len(columns)) + " |",
    ]
    for _, row in df.iterrows():
        values = []
        for col in columns:
            value = row[col]
            if isinstance(value, float):
                values.append(f"{value:.3f}")
            else:
                values.append(str(value))
        lines.append("| " + " | ".join(values) + " |")
    return "\n".join(lines)


def write_report(symptom_df: pd.DataFrame, summary_df: pd.DataFrame, encoder: LabelEncoder) -> None:
    class_counts = symptom_df["pcs_symptom_category"].value_counts().rename_axis("class").reset_index(name="n")
    class_counts["display_name"] = class_counts["class"].map(lambda x: CATEGORY_DISPLAY.get(x, x))
    raw_counts = symptom_df["pcs_symptom_raw"].fillna("None").value_counts().rename_axis("raw_symptom").reset_index(name="n")
    best = summary_df.iloc[0]
    text = f"""# PCS Symptom Occurrence Prediction Task

## Label construction

The original free-text field `PCS症状类型` was converted into structured labels and a merged modeling category. The merged outcome is intended as an additional prediction task and is **not** used as an input feature for the original PCS binary prediction model.

Modeling classes:

{markdown_table(class_counts[["class", "display_name", "n"]])}

Raw symptom text distribution:

{markdown_table(raw_counts)}

## Prediction task

- Task: multiclass prediction of PCS symptom occurrence category.
- Outcome: `pcs_symptom_category`.
- Included classes for modeling: {", ".join(encoder.classes_)}.
- Excluded from modeling: `pcs_unknown`, because only one PCS-positive patient lacked a symptom text label.
- Predictors: the same preoperative/perioperative variables used in the original PCS prediction task.
- Validation: {N_SPLITS}-fold stratified cross-validation.

## Model performance

{markdown_table(summary_df)}

The current top-ranked model by macro F1 is **{best["model"]}** with OOF macro F1 = **{best["oof_macro_f1"]:.3f}**, OOF balanced accuracy = **{best["oof_balanced_accuracy"]:.3f}**, and OOF macro one-vs-rest AUC = **{best["oof_macro_ovr_auc"]:.3f}**.

## Figures

![Class distribution](pcs_symptom_class_distribution.png)

![Confusion matrix](pcs_symptom_confusion_matrix.png)

![One-vs-rest ROC](pcs_symptom_ovr_roc.png)

## Manuscript-ready wording

To further characterize PCS heterogeneity, the free-text PCS symptom field was mapped to structured symptom phenotypes. PCS-negative patients were assigned to `no_pcs`; PCS-positive symptoms were grouped into pain, constipation, and gastrointestinal dysfunction/dyspepsia phenotypes. A secondary multiclass prediction task was then constructed using the same preoperative/perioperative predictors. This task should be interpreted as exploratory because symptom subtype sample sizes are small and external validation is unavailable.
"""
    (OUTPUT / "pcs_symptom_prediction_report.md").write_text(text, encoding="utf-8")


def main() -> None:
    OUTPUT.mkdir(parents=True, exist_ok=True)
    symptom_df = build_symptom_dataset()
    symptom_df.to_csv(OUTPUT / "pcs_symptom_labeled_dataset.csv", index=False, encoding="utf-8-sig")

    label_map = pd.DataFrame(
        [
            {"raw_symptom": raw, "symptom_label": label, "symptom_category": LABEL_TO_CATEGORY[label]}
            for raw, label in RAW_TO_LABEL.items()
        ]
        + [{"raw_symptom": None, "symptom_label": "no_pcs", "symptom_category": "no_pcs"}]
        + [{"raw_symptom": None, "symptom_label": "pcs_unknown", "symptom_category": "pcs_unknown"}]
    )
    label_map.to_csv(OUTPUT / "pcs_symptom_label_mapping.csv", index=False, encoding="utf-8-sig")

    fold_df, summary_df, oof_df, encoder = evaluate_symptom_task(symptom_df)
    fold_df.to_csv(OUTPUT / "pcs_symptom_model_performance_by_fold.csv", index=False, encoding="utf-8-sig")
    summary_df.to_csv(OUTPUT / "pcs_symptom_model_performance_summary.csv", index=False, encoding="utf-8-sig")
    oof_df.to_csv(OUTPUT / "pcs_symptom_oof_predictions.csv", index=False, encoding="utf-8-sig")
    with pd.ExcelWriter(OUTPUT / "pcs_symptom_prediction_outputs.xlsx") as writer:
        symptom_df.to_excel(writer, sheet_name="labeled_dataset", index=False)
        label_map.to_excel(writer, sheet_name="label_mapping", index=False)
        summary_df.to_excel(writer, sheet_name="performance_summary", index=False)
        fold_df.to_excel(writer, sheet_name="fold_metrics", index=False)

    metadata = {
        "task": "multiclass PCS symptom occurrence category prediction",
        "n_rows_labeled_dataset": int(len(symptom_df)),
        "n_rows_modeled": int((symptom_df["pcs_symptom_category"] != "pcs_unknown").sum()),
        "classes": encoder.classes_.tolist(),
        "cv": f"{N_SPLITS}-fold stratified cross-validation",
        "random_state": RANDOM_STATE,
    }
    (OUTPUT / "pcs_symptom_task_metadata.json").write_text(
        json.dumps(metadata, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    write_report(symptom_df, summary_df, encoder)
    print(f"Saved PCS symptom task report to {OUTPUT / 'pcs_symptom_prediction_report.md'}")
    print(summary_df.to_string(index=False))


if __name__ == "__main__":
    main()
